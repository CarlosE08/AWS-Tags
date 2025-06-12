import boto3
import os
from botocore.exceptions import ClientError, ProfileNotFound
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

AWS_REGION = 'us-east-1'  # Modifica según la región

ETIQUETAS_ARCHIVO = "Etiquetas_Comunes.txt"

COLOR_ADDED = "C6EFCE"  # Verde claro para celdas con etiqueta "Added=Yes"

perfil = input("\U0001F464 Por favor, ingresa tu perfil de AWS: ")

try:
    # Verificar si el perfil existe creando una sesión temporal
    boto3.Session(profile_name=perfil)
except ProfileNotFound:
    print(f"❌ El perfil '{perfil}' no se encontró en tu configuración de AWS. Verifica con 'aws configure list-profiles'.")
    exit(1)

# Ruta base de salida
OUTPUT_DIR = os.path.join(os.getcwd(), perfil)
try:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
except Exception as e:
    print(f"❌ Error al crear la carpeta de salida '{OUTPUT_DIR}': {e}")
    exit(1)

# Crea una sesión usando el perfil con SSO
session = boto3.Session(profile_name=perfil, region_name=AWS_REGION)

# Crea un cliente para la API de etiquetado de recursos
CLIENT = session.client('resourcegroupstaggingapi')

# --- Funciones auxiliares ---

def list_all_tagged_resources():
    client = CLIENT
    resources_with_tags = []
    resources_without_tags = []
    paginator = client.get_paginator('get_resources')
    page_iterator = paginator.paginate(ResourcesPerPage=50)
    for page in page_iterator:
        for resource in page['ResourceTagMappingList']:
            arn = resource['ResourceARN']
            tags = resource.get('Tags', [])
            tag_dict = {tag['Key']: tag['Value'] for tag in tags}
            name = tag_dict.get("Name") if tag_dict else None
            if not name:
                name = arn.split("/")[-1] if "/" in arn else arn.split(":")[-1]
            resource_type = infer_resource_type(arn)
            entry = {'ARN': arn, 'Name': name, 'Type': resource_type}
            if tag_dict:
                entry['Tags'] = tag_dict
                resources_with_tags.append(entry)
            else:
                resources_without_tags.append(entry)
    return resources_with_tags, resources_without_tags

def infer_resource_type(arn):
    try:
        parts = arn.split(":")
        service = parts[2]
        resource_section = parts[5] if len(parts) > 5 else ""
        if service == "s3": return "S3"
        elif service == "codecommit": return "CODECOMMIT"
        if "/" in resource_section:
            subtype = resource_section.split("/")[0]
        elif ":" in resource_section:
            subtype = resource_section.split(":")[0]
        else:
            subtype = resource_section
        return f"{service.upper()}::{subtype}" if subtype else service.upper()
    except:
        return "Desconocido"

def autofit_columns(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    cell_length = max(len(str(line)) for line in str(cell.value).split('\n'))
                    if cell_length > max_length:
                        max_length = cell_length
            except: pass
        worksheet.column_dimensions[column_letter].width = max_length + 2

def export_to_excel(resources_with_tags, resources_without_tags):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ConEtiquetas"
    ws1.append(["ARN", "Nombre", "Etiquetas"])
    resource_groups = {}
    for res in resources_with_tags:
        rtype = res['Type']
        resource_groups.setdefault(rtype, []).append(res)
    for rtype in sorted(resource_groups.keys()):
        ws1.append([f"# {rtype}"])
        row_idx = ws1.max_row
        for col in range(1, 4):
            cell = ws1.cell(row=row_idx, column=col)
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[row_idx].height = 30
        for res in resource_groups[rtype]:
            arn = res['ARN']
            name = res['Name']
            tags = res['Tags']
            tag_str = '\n'.join([f"{k}={v}" for k, v in tags.items()])
            ws1.append([arn, name, tag_str])
            cell = ws1.cell(row=ws1.max_row, column=3)
            cell.alignment = Alignment(wrap_text=True)
            if tags.get("Added") == "Yes":
                cell.fill = PatternFill(start_color=COLOR_ADDED, end_color=COLOR_ADDED, fill_type="solid")
    autofit_columns(ws1)

    ws2 = wb.create_sheet("SinEtiquetas")
    ws2.append(["ARN", "Nombre"])
    without_tag_groups = {}
    for res in resources_without_tags:
        rtype = res['Type']
        without_tag_groups.setdefault(rtype, []).append(res)
    for rtype in sorted(without_tag_groups.keys()):
        ws2.append([f"# {rtype}"])
        row_idx = ws2.max_row
        for col in range(1, 3):
            cell = ws2.cell(row=row_idx, column=col)
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[row_idx].height = 30
        for res in without_tag_groups[rtype]:
            ws2.append([res['ARN'], res['Name']])
    autofit_columns(ws2)

    ws3 = wb.create_sheet("Resumen")
    total = len(resources_with_tags) + len(resources_without_tags)
    ws3.append(["Métrica", "Valor"])
    ws3.append(["Total de recursos escaneados", total])
    ws3.append(["Con etiquetas", len(resources_with_tags)])
    ws3.append(["Sin etiquetas", len(resources_without_tags)])
    autofit_columns(ws3)

    output_path = os.path.join(OUTPUT_DIR, f"reporte_etiquetas_aws_{perfil}.xlsx")
    wb.save(output_path)
    print(f"\U0001F4C4 Reporte Excel guardado como '{output_path}'")

def parse_etiquetas_comunes(filepath=ETIQUETAS_ARCHIVO):
    etiquetas = {}
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            if "=" in line:
                key, value = line.split("=", 1)
                etiquetas[key.strip()] = value.strip().strip('"')
    return etiquetas

def es_recurso_etiquetable(arn):
    palabras_reservadas = ["controltower", "awscontroltower", "awscontrolttowermanagedrule", "awsmanaged"]
    arn_lower = arn.lower()
    return not any(palabra in arn_lower for palabra in palabras_reservadas)

def asignar_etiquetas_comunes(etiquetas):
    from collections import defaultdict
    etiquetas_con_added = etiquetas.copy()
    etiquetas_con_added["Added"] = "Yes"
    recursos_etiquetados_exitosos = []
    etiquetados_por_tipo = defaultdict(list)
    client = CLIENT
    arns_to_tag = []
    
    paginator = client.get_paginator('get_resources')
    page_iterator = paginator.paginate(ResourcesPerPage=50)
    
    for page in page_iterator:
        for resource in page['ResourceTagMappingList']:
            arns_to_tag.append(resource['ResourceARN'])
    
    print("⏳ Procesando asignación de etiquetas...")

    for i in range(0, len(arns_to_tag), 20):
        batch = arns_to_tag[i:i + 20]
        filtered_batch = []
        for arn in batch:
            if not es_recurso_etiquetable(arn):
                print(f"⚠️ Recurso no etiquetable (omitido): {arn} — contiene palabra reservada")
                continue
            try:
                tags_response = client.get_resources(ResourceARNList=[arn])
                tag_map = tags_response['ResourceTagMappingList']
                if tag_map and not tag_map[0].get('Tags'):
                    filtered_batch.append(arn)
            except Exception as e:
                print(f"❌ Error al obtener etiquetas de {arn}: {e}")
                continue
        if not filtered_batch:
            continue
        try:
            response = client.tag_resources(ResourceARNList=filtered_batch, Tags=etiquetas_con_added)
            failed_map = response.get("FailedResourcesMap", {})
            for arn in filtered_batch:
                if arn not in failed_map:
                    recursos_etiquetados_exitosos.append(arn)
                    tipo = infer_resource_type(arn)
                    etiquetados_por_tipo[tipo].append(arn)
        except Exception as e:
            print(f"❌ Error al etiquetar lote: {e}")
            continue

    output_txt = os.path.join(OUTPUT_DIR, f"recursos_etiquetados_{perfil}.txt")
    with open(output_txt, "w", encoding="utf-8") as f:
        for tipo in sorted(etiquetados_por_tipo.keys()):
            f.write(f"# {tipo}\n")
            for arn in etiquetados_por_tipo[tipo]:
                f.write(f"- {arn}\n")
                print(f"✅ Etiquetado: {arn}")
            f.write("\n")
    
    print(f"📝 Se exportó la lista de recursos etiquetados en '{output_txt}'")

def eliminar_etiquetas_comunes(etiquetas):
    client = CLIENT
    arns_to_untag = []
    etiquetas_por_recurso = {}
    etiquetas_por_tipo = {}
    keys_to_remove = list(etiquetas.keys()) + ["Added"]
    paginator = client.get_paginator('get_resources')
    page_iterator = paginator.paginate(ResourcesPerPage=50)
    for page in page_iterator:
        for resource in page['ResourceTagMappingList']:
            arn = resource['ResourceARN']
            tags = resource.get("Tags", [])
            if not tags: continue
            tag_dict = {tag['Key']: tag['Value'] for tag in tags}
            if tag_dict.get("Added") != "Yes": continue
            arns_to_untag.append(arn)
            etiquetas_por_recurso[arn] = tag_dict
            tipo = infer_resource_type(arn)
            etiquetas_por_tipo.setdefault(tipo, []).append(arn)
    respaldo_path = os.path.join(OUTPUT_DIR, f"respaldo_etiquetas_eliminadas_{perfil}.txt")
    with open(respaldo_path, "w", encoding="utf-8") as f:
        for tipo in sorted(etiquetas_por_tipo.keys()):
            f.write(f"# {tipo}\n")
            for arn in etiquetas_por_tipo[tipo]:
                f.write(f"- {arn}\n")
                for k, v in etiquetas_por_recurso[arn].items():
                    f.write(f"    {k}={v}\n")
    print(f"📄 Se creó el respaldo en '{respaldo_path}'")
    print("⏳ Eliminando etiquetas en recursos...")
    for i in range(0, len(arns_to_untag), 20):
        batch = arns_to_untag[i:i + 20]
        try:
            client.untag_resources(ResourceARNList=batch, TagKeys=keys_to_remove)
            for arn in batch:
                print(f"🧹 Etiquetas eliminadas de: {arn}")
        except: continue

def restaurar_etiquetas_respaldo(filepath=None):
    import re
    from collections import defaultdict
    client = CLIENT
    if not filepath:
        filepath = os.path.join(OUTPUT_DIR, f"respaldo_etiquetas_eliminadas_{perfil}.txt")
    if not os.path.exists(filepath):
        print("❌ El archivo de respaldo no existe.")
        return
    recurso_tags = defaultdict(dict)
    current_arn = None
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"): continue
            elif line.startswith("- "):
                current_arn = line[2:].strip()
            elif current_arn and "=" in line:
                try:
                    k, v = line.split("=", 1)
                    recurso_tags[current_arn][k.strip()] = v.strip()
                except: continue
    print("⏳ Restaurando etiquetas desde respaldo...")
    for i in range(0, len(recurso_tags), 20):
        batch = list(recurso_tags.keys())[i:i+20]
        for arn in batch:
            try:
                client.tag_resources(ResourceARNList=[arn], Tags=recurso_tags[arn])
                print(f"♻️ Etiquetas restauradas en: {arn}")
            except: continue

def aplicar_merge_a_todos(etiquetas):
    etiquetas_con_added = etiquetas.copy()
    etiquetas_con_added["Added"] = "Yes"
    client = CLIENT
    arns_to_tag = []
    paginator = client.get_paginator('get_resources')
    page_iterator = paginator.paginate(ResourcesPerPage=50)
    for page in page_iterator:
        for resource in page['ResourceTagMappingList']:
            arns_to_tag.append(resource['ResourceARN'])
    print(f"🔖 Aplicando etiquetas a {len(arns_to_tag)} recursos...")
    print("⏳ Aplicando merge de etiquetas...")
    for i in range(0, len(arns_to_tag), 20):
        batch = arns_to_tag[i:i + 20]
        try:
            response = client.tag_resources(ResourceARNList=batch, Tags=etiquetas_con_added)
            failed = response.get("FailedResourcesMap", {})
            for arn in batch:
                if arn not in failed:
                    print(f"✅ Merge de etiquetas aplicado en: {arn}")
        except ClientError as e:
            print(f"❌ Error en batch: {e}")

# --- Menú de ejecución principal ---

if __name__ == "__main__":
    print("\U0001F4CB MENÚ DE OPERACIONES AWS")
    print("1. Escanear recursos y generar Excel")
    print("2. Asignar etiquetas comunes (solo recursos sin etiquetas)")
    print("3. Eliminar etiquetas comunes desde archivo")
    print("4. Asignar etiquetas comunes a todos (merge y sobrescribe claves coincidentes)")
    print("5. Restaurar etiquetas desde respaldo")
    opcion = input("Selecciona una opción (1, 2, 3, 4 o 5): ").strip()
    if opcion == "1":
        with_tags, without_tags = list_all_tagged_resources()
        export_to_excel(with_tags, without_tags)
    elif opcion == "2":
        etiquetas = parse_etiquetas_comunes()
        asignar_etiquetas_comunes(etiquetas)
    elif opcion == "3":
        etiquetas = parse_etiquetas_comunes()
        eliminar_etiquetas_comunes(etiquetas)
    elif opcion == "4":
        etiquetas = parse_etiquetas_comunes()
        aplicar_merge_a_todos(etiquetas)
    elif opcion == "5":
        restaurar_etiquetas_respaldo()
    else:
        print("❌ Opción no válida. Por favor selecciona 1, 2, 3, 4 o 5.")